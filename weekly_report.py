"""
weekly_report.py — Generate DieselUp weekly Excel report
4 sheets:
  1. Summary        — fleet overview, totals
  2. Compliance     — per truck: recommended vs visited stops, savings/losses
  3. Flags          — all flags detail (wrong stop, missed stop, low-stop state)
  4. Fuel By State  — IFTA breakdown by state
"""

import os
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter

# ── Color palette ──────────────────────────────────────────────────────────
BG_DARK    = "060608"
GREEN      = "7FFF5F"
GREEN_DARK = "3A7A2A"
WHITE      = "FAFAFA"
GRAY_DARK  = "1C1C24"
GRAY_MID   = "2A2A35"
GRAY_LIGHT = "3A3A48"
RED        = "FF4545"
AMBER      = "F59E0B"
BLUE       = "3B82F6"

H_BG   = PatternFill("solid", fgColor=GRAY_DARK)
H2_BG  = PatternFill("solid", fgColor=GRAY_MID)
ROW_BG = PatternFill("solid", fgColor="0F0F14")
ALT_BG = PatternFill("solid", fgColor="141420")
TOP_BG = PatternFill("solid", fgColor=GREEN_DARK)

def _font(bold=False, color=WHITE, size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _border():
    s = Side(style="thin", color="2A2A35")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=False)

def _hdr(ws, row, col, value, width=None, color=WHITE, bg=None, size=10, bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=color, size=size)
    c.fill      = bg or H_BG
    c.alignment = _center()
    c.border    = _border()
    if width and isinstance(col, int):
        ws.column_dimensions[get_column_letter(col)].width = width
    return c

def _cell(ws, row, col, value, fmt=None, color=WHITE, bold=False, bg=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=bold, color=color, size=10)
    c.fill      = bg or (ROW_BG if row % 2 == 0 else ALT_BG)
    c.alignment = _center() if align == "center" else _left()
    c.border    = _border()
    if fmt:
        c.number_format = fmt
    return c


# ── MOCK DATA (replace with real DB queries) ─────────────────────────────
def get_real_data(days: int = 7):
    """Pull real data from PostgreSQL for weekly report."""
    from database import (get_flags_for_report, get_compliance_for_report,
                          db_cursor)
    from ifta import IFTA_RATES, HOME_STATE, HOME_STATE_RATE, get_ifta_rate
    from datetime import datetime, timezone, timedelta

    week_start = (datetime.now(timezone.utc) - timedelta(days=days)).strftime("%b %d, %Y")
    week_end   = datetime.now(timezone.utc).strftime("%b %d, %Y")
    since      = datetime.now(timezone.utc) - timedelta(days=days)

    with db_cursor() as cur:
        # Summary stats
        try:
            cur.execute("SELECT COUNT(*) as cnt FROM trucks WHERE is_active=TRUE")
            total_trucks = cur.fetchone()["cnt"]
        except Exception: total_trucks = 0

        try:
            cur.execute("""
                SELECT COUNT(*) as cnt FROM trucks WHERE is_active=TRUE
                AND vehicle_name IN (
                    SELECT DISTINCT vehicle_name FROM fuel_alerts WHERE alerted_at >= %s
                )""", (since,))
            active_trucks = cur.fetchone()["cnt"]
        except Exception: active_trucks = 0

        try:
            cur.execute("SELECT COUNT(*) as cnt FROM fuel_alerts WHERE alerted_at >= %s", (since,))
            total_alerts = cur.fetchone()["cnt"]
        except Exception: total_alerts = 0

        try:
            cur.execute("""
                SELECT
                    COUNT(*) FILTER (WHERE visited=TRUE) as visited,
                    COUNT(*) FILTER (WHERE visited=FALSE) as skipped,
                    COALESCE(SUM(CASE WHEN visited THEN savings_usd ELSE 0 END),0) as savings,
                    COALESCE(SUM(CASE WHEN NOT visited THEN ABS(savings_usd) ELSE 0 END),0) as losses
                FROM stop_visits WHERE visited_at >= %s
            """, (since,))
            sv = cur.fetchone()
            visited_rec = sv["visited"] or 0
            skipped_rec = sv["skipped"] or 0
            total_savings = float(sv["savings"] or 0)
            total_losses  = float(sv["losses"]  or 0)
        except Exception:
            visited_rec = skipped_rec = 0
            total_savings = total_losses = 0.0

        # Override with REAL confirmed losses from driver_flags
        # These are exact losses calculated when driver fueled at a different stop
        try:
            cur.execute("""
                SELECT COALESCE(SUM(savings_lost), 0) as real_lost
                FROM driver_flags
                WHERE flagged_at >= %s
                  AND savings_lost IS NOT NULL AND savings_lost > 0
            """, (since,))
            real_losses = float(cur.fetchone()["real_lost"] or 0)
            if real_losses > 0:
                total_losses = real_losses  # use confirmed real losses
        except Exception:
            pass

        total_refuels  = visited_rec + skipped_rec
        compliance_pct = (visited_rec / total_refuels * 100) if total_refuels else 0
        net_savings    = total_savings - total_losses

        try:
            cur.execute("""
                SELECT AVG(mpg) as avg FROM truck_efficiency WHERE updated_at >= %s
            """, (since,))
            row = cur.fetchone()
            fleet_avg_mpg = round(float(row["avg"]) if row and row["avg"] else 6.5, 1)
        except Exception: fleet_avg_mpg = 6.5

        try:
            cur.execute("""
                SELECT COALESCE(SUM(idle_hours),0) as total FROM truck_efficiency
                WHERE updated_at >= %s
            """, (since,))
            total_idle = round(float(cur.fetchone()["total"] or 0), 0)
        except Exception: total_idle = 0

    summary = {
        "week_start":        week_start,
        "week_end":          week_end,
        "total_trucks":      total_trucks,
        "active_trucks":     active_trucks,
        "total_alerts":      total_alerts,
        "total_refuels":     total_refuels,
        "visited_rec":       visited_rec,
        "skipped_rec":       skipped_rec,
        "compliance_pct":    round(compliance_pct, 1),
        "total_savings_usd": round(total_savings, 2),
        "total_losses_usd":  round(total_losses, 2),
        "net_savings_usd":   round(net_savings, 2),
        "fleet_avg_mpg":     fleet_avg_mpg,
        "total_idle_hrs":    int(total_idle),
        "ifta_est_owed":     0.0,
    }

    # Compliance per truck — merge stop_visits + real losses from driver_flags
    raw_comp = get_compliance_for_report(days)

    # Get real confirmed losses per truck from driver_flags
    real_losses_by_truck = {}
    try:
        with db_cursor() as cur:
            cur.execute("""
                SELECT vehicle_name,
                       COALESCE(SUM(savings_lost), 0) as real_lost,
                       COUNT(*) FILTER (WHERE savings_lost > 0) as confirmed_flags
                FROM driver_flags
                WHERE flagged_at >= %s
                  AND savings_lost IS NOT NULL
                GROUP BY vehicle_name
            """, (since,))
            for r in cur.fetchall():
                real_losses_by_truck[r["vehicle_name"]] = {
                    "real_lost":       float(r["real_lost"] or 0),
                    "confirmed_flags": int(r["confirmed_flags"] or 0),
                }
    except Exception:
        pass

    compliance = []
    for r in raw_comp:
        total       = (r.get("visited") or 0) + (r.get("skipped") or 0)
        truck_name  = r["vehicle_name"]
        real_data   = real_losses_by_truck.get(truck_name, {})
        # Use real confirmed losses if available
        losses      = real_data.get("real_lost") or float(r.get("losses") or 0)
        compliance.append({
            "truck":            truck_name,
            "driver":           "",
            "alerts":           total,
            "visited":          r.get("visited") or 0,
            "skipped":          r.get("skipped") or 0,
            "savings":          float(r.get("savings") or 0),
            "losses":           losses,
            "confirmed_flags":  real_data.get("confirmed_flags", 0),
            "avg_mpg":          6.5,
        })

    # Flags
    raw_flags = get_flags_for_report(days)
    flags = []
    for f in raw_flags:
        ft = f.get("flag_type", "").replace("_", " ").title()
        flags.append({
            "date":         f["flagged_at"].strftime("%b %d %H:%M") if f.get("flagged_at") else "",
            "truck":        f.get("vehicle_name", ""),
            "driver":       "",
            "type":         ft,
            "recommended":  f.get("recommended_stop") or "—",
            "actual":       f.get("actual_stop") or "—",
            "fuel_pct":     int(f.get("fuel_pct") or 0),
            "savings_lost": float(f.get("savings_lost") or 0),
        })

    # IFTA by state — from stop_visits
    ifta_by_state = []
    try:
        with db_cursor() as cur:
            cur.execute("""
                SELECT actual_stop_state as state,
                       COUNT(*) as stops,
                       COALESCE(SUM(gallons_purchased),0) as gallons
                FROM stop_visits
                WHERE visited_at >= %s AND actual_stop_state IS NOT NULL
                GROUP BY actual_stop_state ORDER BY gallons DESC LIMIT 12
            """, (since,))
            rows = cur.fetchall()
            for r in rows:
                st      = r["state"]
                gal     = float(r["gallons"] or 0)
                rate    = get_ifta_rate(st)
                home    = HOME_STATE_RATE
                adj     = round(home - rate, 3)
                total_a = round(adj * gal, 2)
                ifta_by_state.append({
                    "state":       st,
                    "name":        st,
                    "gallons":     int(gal),
                    "rate":        rate,
                    "home_rate":   home,
                    "adj_per_gal": adj,
                    "total_adj":   total_a,
                })
    except Exception: pass

    return summary, compliance, flags, ifta_by_state


# Keep mock data as fallback for testing
def get_mock_data():
    return get_real_data()


def build_report(summary, compliance, flags, ifta_by_state, output_path):
    wb = Workbook()

    # ════════════════════════════════════════════════════════
    # SHEET 1 — SUMMARY
    # ════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "📊 Summary"
    ws1.sheet_view.showGridLines = False

    # Title row
    ws1.merge_cells("A1:F1")
    t = ws1["A1"]
    t.value     = f"DieselUp Weekly Report  |  {summary['week_start']} – {summary['week_end']}"
    t.font      = Font(name="Arial", bold=True, size=14, color=GREEN)
    t.fill      = PatternFill("solid", fgColor=GRAY_DARK)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:F2")
    ws1["A2"].value = ""
    ws1["A2"].fill  = PatternFill("solid", fgColor=GRAY_DARK)
    ws1.row_dimensions[2].height = 8

    # Section: Fleet Overview
    headers = ["Metric", "Value", "", "Metric", "Value", ""]
    for ci, h in enumerate(headers, 1):
        _hdr(ws1, 3, ci, h, bg=H2_BG, color=GREEN if h == "Metric" else WHITE)

    data = [
        ("Total Trucks",          summary["total_trucks"],      "", "Total Alerts Fired",     summary["total_alerts"],   ""),
        ("Active Trucks",         summary["active_trucks"],     "", "Total Refuels Detected",  summary["total_refuels"],  ""),
        ("Visited Recommended",   summary["visited_rec"],       "", "Skipped Recommended",     summary["skipped_rec"],    ""),
        ("Compliance Rate",       f"{summary['compliance_pct']:.1f}%", "", "Fleet Avg MPG",   f"{summary['fleet_avg_mpg']:.1f}", ""),
        ("Total Idle Hours",      summary["total_idle_hrs"],    "", "IFTA Est. Owed",          f"${summary['ifta_est_owed']:,.2f}", ""),
    ]
    for ri, row in enumerate(data, 4):
        for ci, val in enumerate(row, 1):
            bold  = ci in (1, 4)
            color = WHITE if not bold else GREEN
            _cell(ws1, ri, ci, val, color=color, bold=bold)

    ws1.row_dimensions[3].height = 20

    # Savings box
    ws1.merge_cells("A10:F10")
    ws1["A10"].value = ""
    ws1["A10"].fill  = PatternFill("solid", fgColor=GRAY_DARK)

    sav_data = [
        ("A11", "SAVINGS SUMMARY",        True,  GREEN,  "FF0000", 14),
        ("A12", "Total Savings",          False, WHITE,  None,     10),
        ("B12", f"${summary['total_savings_usd']:,.2f}", True, GREEN, None, 12),
        ("A13", "Total Losses (skipped)", False, WHITE,  None,     10),
        ("B13", f"${summary['total_losses_usd']:,.2f}",  True, RED,   None, 12),
        ("A14", "Net Savings",            False, WHITE,  None,     10),
        ("B14", f"${summary['net_savings_usd']:,.2f}",   True, GREEN, None, 14),
    ]
    for ref, val, bold, color, _, size in sav_data:
        c = ws1[ref]
        c.value     = val
        c.font      = Font(name="Arial", bold=bold, color=color, size=size)
        c.fill      = PatternFill("solid", fgColor=GRAY_MID)
        c.alignment = _left()

    for col, width in zip("ABCDEF", [28, 18, 4, 28, 18, 4]):
        ws1.column_dimensions[col].width = width
    for r in range(4, 10):
        ws1.row_dimensions[r].height = 22

    # ════════════════════════════════════════════════════════
    # SHEET 2 — COMPLIANCE
    # ════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("✅ Compliance")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:J1")
    t2 = ws2["A1"]
    t2.value     = "Driver Compliance Report — Recommended Stop Tracking"
    t2.font      = Font(name="Arial", bold=True, size=13, color=GREEN)
    t2.fill      = PatternFill("solid", fgColor=GRAY_DARK)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    cols = ["Truck #", "Driver Name", "Alerts\nSent", "Visited\nRec. Stop",
            "Skipped\nRec. Stop", "Compliance\n%", "Savings\n($)",
            "Real Losses\n($)", "Confirmed\nFlags", "Net\n($)", "Avg\nMPG"]
    widths = [10, 22, 10, 12, 12, 12, 12, 14, 12, 12, 10]
    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        _hdr(ws2, 2, ci, h, width=w, bg=H2_BG, color=GREEN)

    for ri, row in enumerate(compliance, 3):
        comp       = row["visited"] / row["alerts"] * 100 if row["alerts"] else 0
        net        = row["savings"] - row["losses"]
        comp_color = GREEN if comp >= 70 else (AMBER if comp >= 50 else RED)
        net_color  = GREEN if net >= 0 else RED
        conf_flags = row.get("confirmed_flags", 0)

        _cell(ws2, ri, 1,  row["truck"],              bold=True, color=WHITE)
        _cell(ws2, ri, 2,  row["driver"],             color=WHITE)
        _cell(ws2, ri, 3,  row["alerts"],             align="center", color=WHITE)
        _cell(ws2, ri, 4,  row["visited"],            align="center", color=GREEN)
        _cell(ws2, ri, 5,  row["skipped"],            align="center", color=RED if row["skipped"] > 0 else WHITE)
        _cell(ws2, ri, 6,  f"{comp:.0f}%",            align="center", color=comp_color, bold=True)
        _cell(ws2, ri, 7,  row["savings"],            fmt='$#,##0.00', align="center", color=GREEN)
        _cell(ws2, ri, 8,  row["losses"],             fmt='$#,##0.00', align="center",
              color=RED if row["losses"] > 0 else WHITE, bold=row["losses"] > 0)
        _cell(ws2, ri, 9,  conf_flags,                align="center",
              color=RED if conf_flags > 0 else WHITE)
        _cell(ws2, ri, 10, net,                       fmt='$#,##0.00', align="center",
              color=net_color, bold=True)
        _cell(ws2, ri, 11, f"{row['avg_mpg']:.1f}",  align="center", color=WHITE)

    # Totals row
    tr = len(compliance) + 3
    tot_alerts  = sum(r["alerts"]  for r in compliance)
    tot_visited = sum(r["visited"] for r in compliance)
    tot_skipped = sum(r["skipped"] for r in compliance)
    tot_savings = sum(r["savings"] for r in compliance)
    tot_losses  = sum(r["losses"]  for r in compliance)
    tot_conf    = sum(r.get("confirmed_flags", 0) for r in compliance)
    tot_net     = tot_savings - tot_losses
    tot_comp    = tot_visited / tot_alerts * 100 if tot_alerts else 0

    total_fill = PatternFill("solid", fgColor=GREEN_DARK)
    totals = [("TOTAL", WHITE, True), ("", WHITE, False),
              (tot_alerts, WHITE, True), (tot_visited, GREEN, True),
              (tot_skipped, RED if tot_skipped else WHITE, True),
              (f"{tot_comp:.0f}%", GREEN if tot_comp >= 70 else AMBER, True),
              (tot_savings, GREEN, True),
              (tot_losses, RED if tot_losses else WHITE, True),
              (tot_conf, RED if tot_conf else WHITE, True),
              (tot_net, GREEN if tot_net >= 0 else RED, True),
              ("", WHITE, False)]
    for ci, (val, color, bold) in enumerate(totals, 1):
        c = ws2.cell(row=tr, column=ci, value=val)
        c.font      = Font(name="Arial", bold=bold, color=color, size=10)
        c.fill      = total_fill
        c.alignment = _center()
        c.border    = _border()
        if ci in (7, 8, 10) and isinstance(val, float):
            c.number_format = '$#,##0.00'

    ws2.row_dimensions[1].height = 28
    ws2.row_dimensions[2].height = 36
    for r in range(3, tr + 1):
        ws2.row_dimensions[r].height = 22

    # ════════════════════════════════════════════════════════
    # SHEET 3 — FLAGS
    # ════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("🚩 Flags")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:H1")
    t3 = ws3["A1"]
    t3.value     = "Driver Flags — Deviations from Recommended Fuel Stops"
    t3.font      = Font(name="Arial", bold=True, size=13, color=WHITE)
    t3.fill      = PatternFill("solid", fgColor="8B0000")
    t3.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    flag_cols = ["Date/Time", "Truck #", "Driver", "Flag Type",
                 "Recommended Stop", "Actual Stop / Issue",
                 "Fuel %", "Savings Lost ($)"]
    flag_widths = [16, 10, 22, 18, 32, 32, 10, 16]
    for ci, (h, w) in enumerate(zip(flag_cols, flag_widths), 1):
        _hdr(ws3, 2, ci, h, width=w, bg=H2_BG, color=GREEN)

    flag_colors = {
        "Wrong Stop":     RED,
        "Missed Stop":    AMBER,
        "Low-Stop State": "FF8800",
    }
    for ri, f in enumerate(flags, 3):
        ft_color = flag_colors.get(f["type"], WHITE)
        _cell(ws3, ri, 1, f["date"],              color=WHITE)
        _cell(ws3, ri, 2, f["truck"],             bold=True, color=WHITE)
        _cell(ws3, ri, 3, f["driver"],            color=WHITE)
        _cell(ws3, ri, 4, f["type"],              bold=True, color=ft_color)
        _cell(ws3, ri, 5, f["recommended"],       color=WHITE)
        _cell(ws3, ri, 6, f["actual"],            color=RED if f["actual"] != "—" else AMBER)
        _cell(ws3, ri, 7, f"{f['fuel_pct']}%",   align="center", color=AMBER if f["fuel_pct"] < 30 else WHITE)
        _cell(ws3, ri, 8, f["savings_lost"] if f["savings_lost"] else "—",
              fmt='$#,##0.00' if f["savings_lost"] else None,
              align="center",
              color=RED if f["savings_lost"] > 0 else WHITE)

    # Flag summary at bottom
    sr = len(flags) + 4
    ws3.merge_cells(f"A{sr}:H{sr}")
    ws3[f"A{sr}"].value     = "FLAG SUMMARY"
    ws3[f"A{sr}"].font      = Font(name="Arial", bold=True, color=GREEN, size=11)
    ws3[f"A{sr}"].fill      = PatternFill("solid", fgColor=GRAY_DARK)
    ws3[f"A{sr}"].alignment = _left()

    from collections import Counter
    flag_counts  = Counter(f["type"] for f in flags)
    total_lost   = sum(f["savings_lost"] for f in flags)
    sr += 1
    for ft, cnt in flag_counts.items():
        ws3[f"A{sr}"].value = f"  {ft}:"
        ws3[f"A{sr}"].font  = Font(name="Arial", color=flag_colors.get(ft, WHITE), size=10)
        ws3[f"B{sr}"].value = f"{cnt} incidents"
        ws3[f"B{sr}"].font  = Font(name="Arial", bold=True, color=WHITE, size=10)
        ws3[f"A{sr}"].fill = ws3[f"B{sr}"].fill = PatternFill("solid", fgColor=GRAY_MID)
        sr += 1

    ws3[f"A{sr}"].value = "  Total Savings Lost:"
    ws3[f"B{sr}"].value = f"${total_lost:,.2f}"
    ws3[f"A{sr}"].font  = Font(name="Arial", color=WHITE, size=10)
    ws3[f"B{sr}"].font  = Font(name="Arial", bold=True, color=RED, size=11)
    ws3[f"A{sr}"].fill  = ws3[f"B{sr}"].fill = PatternFill("solid", fgColor=GRAY_MID)

    ws3.row_dimensions[1].height = 28
    ws3.row_dimensions[2].height = 36
    for r in range(3, len(flags) + 3):
        ws3.row_dimensions[r].height = 22

    # ════════════════════════════════════════════════════════
    # SHEET 4 — FUEL BY STATE (IFTA)
    # ════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("📋 IFTA by State")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:G1")
    t4 = ws4["A1"]
    t4.value     = "IFTA Fuel Settlement Estimate — by State"
    t4.font      = Font(name="Arial", bold=True, size=13, color=GREEN)
    t4.fill      = PatternFill("solid", fgColor=GRAY_DARK)
    t4.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    ifta_cols = ["State", "State Name", "Gallons\nPurchased",
                 "State Rate\n($/gal)", "Home Rate\n($/gal)",
                 "Adj Per\nGallon", "Total IFTA\nAdj ($)"]
    ifta_widths = [8, 18, 14, 14, 14, 14, 16]
    for ci, (h, w) in enumerate(zip(ifta_cols, ifta_widths), 1):
        _hdr(ws4, 2, ci, h, width=w, bg=H2_BG, color=GREEN)

    for ri, row in enumerate(ifta_by_state, 3):
        adj_color = RED if row["total_adj"] > 0 else GREEN
        _cell(ws4, ri, 1, row["state"],        bold=True, color=WHITE, align="center")
        _cell(ws4, ri, 2, row["name"],         color=WHITE)
        _cell(ws4, ri, 3, row["gallons"],      fmt='#,##0', align="center", color=WHITE)
        _cell(ws4, ri, 4, row["rate"],         fmt='$#,##0.000', align="center", color=WHITE)
        _cell(ws4, ri, 5, row["home_rate"],    fmt='$#,##0.000', align="center", color=WHITE)
        _cell(ws4, ri, 6, row["adj_per_gal"],  fmt='$#,##0.000', align="center",
              color=RED if row["adj_per_gal"] > 0 else GREEN, bold=True)
        _cell(ws4, ri, 7, row["total_adj"],    fmt='$#,##0.00', align="center",
              color=adj_color, bold=True)

    # Totals
    tr4 = len(ifta_by_state) + 3
    tot_gal  = sum(r["gallons"]   for r in ifta_by_state)
    tot_ifta = sum(r["total_adj"] for r in ifta_by_state)
    tf = PatternFill("solid", fgColor=GREEN_DARK)
    for ci, (val, fmt, color) in enumerate([
        ("TOTAL", None, WHITE), ("", None, WHITE),
        (tot_gal, '#,##0', WHITE), ("", None, WHITE),
        ("", None, WHITE), ("", None, WHITE),
        (tot_ifta, '$#,##0.00', RED if tot_ifta > 0 else GREEN)
    ], 1):
        c = ws4.cell(row=tr4, column=ci, value=val)
        c.font      = Font(name="Arial", bold=True, color=color, size=10)
        c.fill      = tf
        c.alignment = _center()
        c.border    = _border()
        if fmt and val != "":
            c.number_format = fmt

    # IFTA note
    nr = tr4 + 2
    ws4[f"A{nr}"].value = (
        f"Positive adjustment = you OWE this amount to home state at settlement.\n"
        f"Negative adjustment = you will receive a CREDIT from home state.\n"
        f"Total estimated IFTA settlement: ${tot_ifta:+,.2f}"
    )
    ws4[f"A{nr}"].font      = Font(name="Arial", italic=True, color=AMBER, size=9)
    ws4[f"A{nr}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws4.merge_cells(f"A{nr}:G{nr}")
    ws4.row_dimensions[nr].height = 52

    ws4.row_dimensions[1].height = 28
    ws4.row_dimensions[2].height = 36
    for r in range(3, tr4 + 1):
        ws4.row_dimensions[r].height = 22

    # ── Global tab colors ──
    ws1.sheet_properties.tabColor = "3A7A2A"
    ws2.sheet_properties.tabColor = "3B82F6"
    ws3.sheet_properties.tabColor = "8B0000"
    ws4.sheet_properties.tabColor = "F59E0B"

    wb.save(output_path)
    print(f"✅ Saved: {output_path}")


if __name__ == "__main__":
    summary, compliance, flags, ifta_by_state = get_mock_data()
    build_report(summary, compliance, flags, ifta_by_state,
                 "/mnt/user-data/outputs/DieselUp_Weekly_Report.xlsx")


# ════════════════════════════════════════════════════════════════════════════
# PER-TRUCK WEEKLY REPORT
# One sheet per truck showing: fuel stops, compliance, flags, savings/losses
# ════════════════════════════════════════════════════════════════════════════

def get_truck_data_from_db(days: int = 7) -> dict:
    """Pull per-truck fuel stops and flags from DB."""
    from database import db_cursor
    from datetime import datetime, timezone, timedelta
    since = datetime.now(timezone.utc) - timedelta(days=days)

    trucks = {}
    try:
        with db_cursor() as cur:
            # Get all active trucks
            cur.execute("SELECT vehicle_name FROM trucks WHERE is_active=TRUE ORDER BY vehicle_name")
            truck_names = [r["vehicle_name"] for r in cur.fetchall()]

        for vname in truck_names:
            with db_cursor() as cur:
                # Fuel stops from stop_visits
                try:
                    cur.execute("""
                        SELECT sv.visited_at as date,
                               sv.actual_stop_name as station,
                               sv.actual_stop_state as state,
                               sv.visited,
                               sv.savings_usd,
                               fa.fuel_pct,
                               fa.alert_type
                        FROM stop_visits sv
                        LEFT JOIN fuel_alerts fa ON sv.alert_id = fa.id
                        WHERE sv.vehicle_name = %s AND sv.visited_at >= %s
                        ORDER BY sv.visited_at
                    """, (vname, since))
                    stops_raw = cur.fetchall()
                except Exception:
                    stops_raw = []

                # Flags
                try:
                    cur.execute("""
                        SELECT flag_type, details, recommended_stop,
                               actual_stop, fuel_pct, savings_lost, flagged_at
                        FROM driver_flags
                        WHERE vehicle_name = %s AND flagged_at >= %s
                        ORDER BY flagged_at
                    """, (vname, since))
                    flags_raw = cur.fetchall()
                except Exception:
                    flags_raw = []

            if not stops_raw and not flags_raw:
                continue

            stops = []
            for r in stops_raw:
                stops.append({
                    "date":        r["date"].strftime("%b %d %H:%M") if r.get("date") else "",
                    "trip":        "",
                    "station":     r.get("station") or "Unknown",
                    "city":        "",
                    "state":       r.get("state") or "",
                    "recommended": bool(r.get("visited")),
                    "retail":      0,
                    "card":        0,
                    "net":         0,
                    "gallons":     0,
                    "pump_paid":   float(r.get("savings_usd") or 0),
                    "net_cost":    0,
                })

            flags = []
            for f in flags_raw:
                ft = (f.get("flag_type") or "").replace("_", " ").title()
                flags.append({
                    "date":        f["flagged_at"].strftime("%b %d %H:%M") if f.get("flagged_at") else "",
                    "type":        ft,
                    "recommended": f.get("recommended_stop") or "—",
                    "actual":      f.get("actual_stop") or "—",
                    "fuel_pct":    int(f.get("fuel_pct") or 0),
                    "loss":        float(f.get("savings_lost") or 0),
                })

            visited  = sum(1 for s in stops if s["recommended"])
            total_sv = len(stops)
            trucks[vname] = {
                "driver": vname,
                "fuel_stops": stops,
                "flags": flags,
                "stats": {
                    "trips":          0,
                    "total_gallons":  0,
                    "total_pump":     sum(s["pump_paid"] for s in stops),
                    "total_net":      sum(s["net_cost"] for s in stops),
                    "savings":        0,
                    "losses":         sum(f["loss"] for f in flags if f["loss"] > 0),
                    "compliance_pct": int(visited / total_sv * 100) if total_sv else 0,
                    "avg_mpg":        6.5,
                },
            }
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"get_truck_data_from_db failed: {e}")

    return trucks

