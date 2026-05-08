"""
pilotfj_importer.py — Parse Pilot/Flying J daily pricing Excel into fuel_stops.

File format (Better Of Pricing Report):
  Rows 0-4: metadata/title rows (skipped automatically)
  Row 5:    headers — Site, City, ST, Prod, ..., Your Price, Savings Total
  Row 6+:   data — one row per store, DSL product only

Routing uses `Your Price` (your negotiated fleet rate) as the card price.
Coordinates come from the pilotfj_locations table (seeded once).
"""

import logging

log = logging.getLogger(__name__)

_HEADER_MARKERS  = ("Site", "City", "ST", "Prod")
_DIESEL_KEYWORDS = {"DSL", "DIESEL", "ULSD", "#2", "D2", "CLEAR"}


def _is_diesel(prod: str) -> bool:
    p = prod.upper().strip()
    return any(k in p for k in _DIESEL_KEYWORDS)


def _find_header_row_xls(ws, max_scan: int = 15) -> int | None:
    for r in range(max_scan):
        row = [str(ws.cell_value(r, c)).strip() for c in range(min(ws.ncols, 10))]
        if all(m in row for m in _HEADER_MARKERS):
            return r
    return None


def _find_header_row_xlsx(ws, max_scan: int = 15) -> int | None:
    for r_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True)):
        cells = [str(c).strip() if c is not None else "" for c in row[:10]]
        if all(m in cells for m in _HEADER_MARKERS):
            return r_idx
    return None


def _to_float(val) -> float | None:
    try:
        f = float(val)
        return f if f > 0 else None
    except (TypeError, ValueError):
        return None


def _pad_store(site: str) -> str:
    s = str(site).strip().rstrip(".0")
    return s.zfill(3) if s.isdigit() else s


def _parse_xls(file_bytes: bytes) -> tuple[list[dict], str]:
    import xlrd
    try:
        wb = xlrd.open_workbook(file_contents=file_bytes)
        ws = wb.sheet_by_index(0)
    except Exception as e:
        return [], f"❌ Could not open .xls file: {e}"

    hr = _find_header_row_xls(ws)
    if hr is None:
        return [], "❌ Header row not found — expected Site, City, ST, Prod columns."

    headers = [str(ws.cell_value(hr, c)).strip() for c in range(ws.ncols)]

    def col(name: str) -> int | None:
        return headers.index(name) if name in headers else None

    # Map columns — handle slight spacing variations in header names
    c_site    = col("Site")
    c_city    = col("City")
    c_st      = col("ST")
    c_prod    = col("Prod")
    c_price   = col("Your Price")
    c_retail  = col("Retail Price")
    c_savings = col("Savings Total")
    c_cost    = col("Cost")
    # State tax header may have extra space
    c_stax = next(
        (i for i, h in enumerate(headers) if "state tax" in h.lower()),
        None,
    )

    if c_site is None or c_price is None:
        return [], "❌ Missing required columns: Site and/or Your Price."

    records          = []
    skipped_no_price = 0
    skipped_non_dsl  = 0

    for r in range(hr + 1, ws.nrows):
        raw_site = str(ws.cell_value(r, c_site)).strip() if c_site is not None else ""
        if not raw_site or raw_site in ("0", ""):
            continue

        prod = str(ws.cell_value(r, c_prod)).strip() if c_prod is not None else "DSL"
        if prod and not _is_diesel(prod):
            skipped_non_dsl += 1
            continue

        your_price = _to_float(ws.cell_value(r, c_price))
        if your_price is None:
            skipped_no_price += 1
            continue

        records.append({
            "store_number":  _pad_store(raw_site),
            "city":          str(ws.cell_value(r, c_city)).strip()  if c_city  is not None else "",
            "state":         str(ws.cell_value(r, c_st)).strip()    if c_st    is not None else "",
            "your_price":    round(your_price, 4),
            "retail_price":  _to_float(ws.cell_value(r, c_retail))  if c_retail  is not None else None,
            "savings_total": _to_float(ws.cell_value(r, c_savings)) if c_savings is not None else None,
            "cost":          _to_float(ws.cell_value(r, c_cost))    if c_cost    is not None else None,
            "state_tax":     _to_float(ws.cell_value(r, c_stax))    if c_stax    is not None else None,
        })

    summary = (
        f"{len(records)} diesel stops parsed"
        + (f", {skipped_non_dsl} non-diesel skipped"   if skipped_non_dsl  else "")
        + (f", {skipped_no_price} missing price skipped" if skipped_no_price else "")
    )
    return records, summary


def _parse_xlsx(file_bytes: bytes) -> tuple[list[dict], str]:
    import openpyxl, io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return [], f"❌ Could not open .xlsx file: {e}"

    all_rows = list(ws.iter_rows(values_only=True))
    hr = _find_header_row_xlsx(ws)
    if hr is None:
        return [], "❌ Header row not found — expected Site, City, ST, Prod columns."

    headers = [str(c).strip() if c is not None else "" for c in all_rows[hr]]

    def col(name: str) -> int | None:
        return headers.index(name) if name in headers else None

    c_site    = col("Site")
    c_city    = col("City")
    c_st      = col("ST")
    c_prod    = col("Prod")
    c_price   = col("Your Price")
    c_retail  = col("Retail Price")
    c_savings = col("Savings Total")
    c_cost    = col("Cost")
    c_stax    = next(
        (i for i, h in enumerate(headers) if "state tax" in h.lower()),
        None,
    )

    if c_site is None or c_price is None:
        return [], "❌ Missing required columns: Site and/or Your Price."

    records          = []
    skipped_no_price = 0
    skipped_non_dsl  = 0

    for row in all_rows[hr + 1:]:
        raw_site = str(row[c_site]).strip() if c_site is not None and row[c_site] is not None else ""
        if not raw_site or raw_site in ("0", "None"):
            continue

        prod = str(row[c_prod]).strip() if c_prod is not None and row[c_prod] is not None else "DSL"
        if prod and not _is_diesel(prod):
            skipped_non_dsl += 1
            continue

        your_price = _to_float(row[c_price] if c_price is not None else None)
        if your_price is None:
            skipped_no_price += 1
            continue

        records.append({
            "store_number":  _pad_store(raw_site),
            "city":          str(row[c_city]).strip()  if c_city  is not None and row[c_city]  else "",
            "state":         str(row[c_st]).strip()    if c_st    is not None and row[c_st]    else "",
            "your_price":    round(your_price, 4),
            "retail_price":  _to_float(row[c_retail])  if c_retail  is not None else None,
            "savings_total": _to_float(row[c_savings]) if c_savings is not None else None,
            "cost":          _to_float(row[c_cost])    if c_cost    is not None else None,
            "state_tax":     _to_float(row[c_stax])    if c_stax    is not None else None,
        })

    summary = (
        f"{len(records)} diesel stops parsed"
        + (f", {skipped_non_dsl} non-diesel skipped"    if skipped_non_dsl  else "")
        + (f", {skipped_no_price} missing price skipped" if skipped_no_price else "")
    )
    return records, summary


def parse_pilotfj_excel(file_bytes: bytes, filename: str = "") -> tuple[list[dict], str]:
    """
    Parse a Pilot/FJ pricing file (.xls or .xlsx).
    Returns (records, summary_message).
    Each record: {store_number, city, state, your_price, retail_price, ...}
    """
    fname = (filename or "").lower()
    if fname.endswith(".xlsx"):
        return _parse_xlsx(file_bytes)
    # Default: try xls first, fall back to xlsx
    records, summary = _parse_xls(file_bytes)
    if not records and "Could not open" not in summary:
        records, summary = _parse_xlsx(file_bytes)
    return records, summary


def import_pilotfj_excel(file_bytes: bytes, filename: str = "") -> tuple[int, str]:
    """
    Parse Pilot/FJ pricing Excel and load into fuel_stops.
    Joins with pilotfj_locations to get coordinates.
    Replaces all existing fuel_stops data.
    Returns (count_loaded, message).
    """
    from database import db_cursor

    records, parse_summary = parse_pilotfj_excel(file_bytes, filename)
    if not records:
        return 0, parse_summary

    # Look up coordinates from pilotfj_locations
    with db_cursor() as cur:
        cur.execute(
            "SELECT store_number, store_name, address, latitude, longitude "
            "FROM pilotfj_locations"
        )
        loc_rows = cur.fetchall()

    loc_map = {str(row["store_number"]).zfill(3): dict(row) for row in loc_rows}

    loaded     = []
    no_coords  = []

    for r in records:
        loc = loc_map.get(r["store_number"])
        if not loc:
            no_coords.append(r["store_number"])
            continue
        lat = loc.get("latitude")
        lng = loc.get("longitude")
        if not lat or not lng:
            no_coords.append(r["store_number"])
            continue
        loaded.append({
            "station_name":     loc["store_name"] or f"Pilot/FJ #{r['store_number']}",
            "address":          loc.get("address") or "",
            "city":             r["city"],
            "state":            r["state"],
            "longitude":        float(lng),
            "latitude":         float(lat),
            "retail_price":     r["retail_price"],
            "discounted_price": r["your_price"],
        })

    if not loaded:
        return 0, (
            f"⚠️ {parse_summary}\n"
            f"❌ 0 stops loaded — no coordinates found for any store.\n"
            f"Upload your Pilot/FJ locations file first (/seedlocations command)."
        )

    with db_cursor() as cur:
        cur.execute("TRUNCATE TABLE fuel_stops RESTART IDENTITY")
        cur.executemany("""
            INSERT INTO fuel_stops
                (station_name, address, city, state, longitude, latitude,
                 retail_price, discounted_price, price_updated)
            VALUES
                (%(station_name)s, %(address)s, %(city)s, %(state)s,
                 %(longitude)s, %(latitude)s,
                 %(retail_price)s, %(discounted_price)s, NOW())
        """, loaded)

    lines = [
        f"✅ *Pilot/Flying J prices updated*",
        f"⛽ {len(loaded)} stops loaded",
        f"💳 Using Your Price (fleet rate) for all routing",
    ]
    if no_coords:
        lines.append(
            f"⚠️ {len(no_coords)} stores skipped — no coordinates yet "
            f"(upload locations file to add them)"
        )
    lines.append(f"📊 {parse_summary}")
    return len(loaded), "\n".join(lines)
