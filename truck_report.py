"""
truck_report.py — Per-truck weekly Excel report

One sheet per truck:
- Every fuel stop this week with station, state, card price, savings, followed?
- Every flag with type, recommended vs actual, fuel %, savings lost
- Summary stats: compliance %, savings, losses, net
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone, timedelta

DARK  = "060608"; CARD  = "101014"; CARD2 = "1C1C24"
GREEN = "7FFF5F"; GDRK  = "3A7A2A"; RED   = "FF4545"
AMBER = "F59E0B"; WHITE = "F8F8FF"; MUTED = "505060"
FLAG_COLORS = {"Wrong Stop": RED, "Missed Stop": AMBER, "Low-Stop State": "FF8800"}

def _fill(c):      return PatternFill("solid", fgColor=c)
def _border():
    s = Side(style="thin", color="2A2A35")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="left"): return Alignment(horizontal=h, vertical="center", wrap_text=False)
def _hdr(ws, r, c, v, w=None, color=WHITE, bg=CARD2, bold=True):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=bold, color=color, size=10)
    cell.fill = _fill(bg); cell.alignment = _align("center"); cell.border = _border()
    if w: ws.column_dimensions[get_column_letter(c)].width = w
    return cell
def _cell(ws, r, c, v, fmt=None, color=WHITE, bold=False, center=False):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=bold, color=color, size=10)
    cell.fill = _fill(DARK if r % 2 == 0 else "141420")
    cell.alignment = _align("center" if center else "left"); cell.border = _border()
    if fmt: cell.number_format = fmt
    return cell
def _sec(ws, row, text, ncols, color=GREEN):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(name="Arial", bold=True, color=color, size=11)
    c.fill = _fill(CARD2); c.alignment = _align("left")
    ws.row_dimensions[row].height = 22


def get_per_truck_data(days: int = 7) -> dict:
    from database import db_cursor
    since = datetime.now(timezone.utc) - timedelta(days=days)
    trucks = {}
    try:
        with db_cursor() as cur:
            cur.execute("SELECT vehicle_name FROM trucks WHERE is_active=TRUE ORDER BY vehicle_name")
            names = [r["vehicle_name"] for r in cur.fetchall()]
        for vname in names:
            with db_cursor() as cur:
                try:
                    cur.execute("""
                        SELECT sv.visited_at, sv.actual_stop_name, sv.actual_stop_state,
                               sv.recommended_stop_name, sv.visited, sv.savings_usd,
                               fa.fuel_pct, fa.best_stop_price
                        FROM stop_visits sv
                        LEFT JOIN fuel_alerts fa ON sv.alert_id = fa.id
                        WHERE sv.vehicle_name = %s AND sv.visited_at >= %s
                        ORDER BY sv.visited_at
                    """, (vname, since))
                    stops_raw = cur.fetchall()
                except Exception:
                    stops_raw = []
                try:
                    cur.execute("""
                        SELECT flag_type, recommended_stop, actual_stop,
                               fuel_pct, savings_lost, flagged_at
                        FROM driver_flags
                        WHERE vehicle_name = %s AND flagged_at >= %s
                        ORDER BY flagged_at
                    """, (vname, since))
                    flags_raw = cur.fetchall()
                except Exception:
                    flags_raw = []
            if not stops_raw and not flags_raw:
                continue
            stops = [{
                "date":        r["visited_at"].strftime("%b %d %H:%M") if r.get("visited_at") else "",
                "station":     r.get("actual_stop_name") or "Unknown",
                "state":       r.get("actual_stop_state") or "",
                "recommended": r.get("recommended_stop_name") or "",
                "followed":    bool(r.get("visited")),
                "fuel_pct":    float(r.get("fuel_pct") or 0),
                "card_price":  float(r.get("best_stop_price") or 0),
                "savings":     float(r.get("savings_usd") or 0),
            } for r in stops_raw]
            flags = [{
                "date":        f["flagged_at"].strftime("%b %d %H:%M") if f.get("flagged_at") else "",
                "type":        (f.get("flag_type") or "").replace("_", " ").title(),
                "recommended": f.get("recommended_stop") or "---",
                "actual":      f.get("actual_stop") or "---",
                "fuel_pct":    int(f.get("fuel_pct") or 0),
                "loss":        float(f.get("savings_lost") or 0),
            } for f in flags_raw]
            visited = sum(1 for s in stops if s["followed"])
            total_sv = len(stops)
            savings = sum(s["savings"] for s in stops if s["followed"])
            losses  = sum(f["loss"] for f in flags if f["loss"] > 0)
            trucks[vname] = {
                "driver": vname, "fuel_stops": stops, "flags": flags,
                "stats": {
                    "total_stops":    total_sv,
                    "visited":        visited,
                    "skipped":        total_sv - visited,
                    "compliance_pct": int(visited / total_sv * 100) if total_sv else 0,
                    "total_savings":  round(savings, 2),
                    "total_losses":   round(losses, 2),
                    "net_savings":    round(savings - losses, 2),
                    "flag_count":     len(flags),
                },
            }
    except Exception as e:
        import logging; logging.getLogger(__name__).error(f"per_truck_data: {e}")
    return trucks


def build_truck_report(output_path: str, days: int = 7) -> None:
    week_end   = datetime.now(timezone.utc)
    week_start = week_end - timedelta(days=days)
    period     = week_start.strftime("%b %d") + " - " + week_end.strftime("%b %d, %Y")
    trucks     = get_per_truck_data(days)
    wb         = Workbook()

    # ── Fleet Summary ────────────────────────────────────────────────────────
    ws_s = wb.active; ws_s.title = "Fleet Summary"
    ws_s.sheet_view.showGridLines = False; ws_s.sheet_properties.tabColor = GDRK
    ws_s.merge_cells("A1:J1")
    t = ws_s["A1"]; t.value = "DieselUp Truck Report  |  " + period
    t.font = Font(name="Arial", bold=True, color=GREEN, size=13)
    t.fill = _fill(CARD2); t.alignment = _align("center")
    ws_s.row_dimensions[1].height = 30
    sh = [("Truck",10),("Driver",20),("Stops",8),("Followed",9),("Skipped",9),
          ("Compliance",12),("Savings ($)",13),("Losses ($)",12),("Net ($)",13),("Flags",7)]
    for ci,(h,w) in enumerate(sh,1): _hdr(ws_s,2,ci,h,w=w,color=GREEN)
    ws_s.row_dimensions[2].height = 28
    for ri,(vname,td) in enumerate(trucks.items(),3):
        s=td["stats"]
        cc = GREEN if s["compliance_pct"]>=80 else (AMBER if s["compliance_pct"]>=60 else RED)
        nc = GREEN if s["net_savings"]>=0 else RED
        _cell(ws_s,ri,1,vname,bold=True); _cell(ws_s,ri,2,td["driver"])
        _cell(ws_s,ri,3,s["total_stops"],center=True)
        _cell(ws_s,ri,4,s["visited"],color=GREEN,center=True)
        _cell(ws_s,ri,5,s["skipped"],color=RED if s["skipped"] else WHITE,center=True)
        _cell(ws_s,ri,6,str(s["compliance_pct"])+"%",color=cc,bold=True,center=True)
        _cell(ws_s,ri,7,s["total_savings"],fmt="$#,##0.00",color=GREEN,center=True)
        _cell(ws_s,ri,8,s["total_losses"],fmt="$#,##0.00",color=RED if s["total_losses"] else WHITE,center=True)
        _cell(ws_s,ri,9,s["net_savings"],fmt="$#,##0.00",color=nc,bold=True,center=True)
        _cell(ws_s,ri,10,s["flag_count"],color=RED if s["flag_count"] else WHITE,bold=s["flag_count"]>0,center=True)
        ws_s.row_dimensions[ri].height = 20

    # ── Per-Truck Sheets ─────────────────────────────────────────────────────
    for vname,td in trucks.items():
        s  = td["stats"]
        ws = wb.create_sheet(vname[:31])
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = (GDRK if s["compliance_pct"]==100
                                        else AMBER if s["compliance_pct"]>=60 else "8B0000")
        ws.merge_cells("A1:H1")
        t = ws["A1"]; t.value = "Truck " + vname + "  |  " + period
        t.font = Font(name="Arial",bold=True,color=GREEN,size=13)
        t.fill = _fill(CARD2); t.alignment = _align("center")
        ws.row_dimensions[1].height = 28

        # Stats bar
        slabels = ["Stops","Followed","Skipped","Compliance","Savings","Losses","Net","Flags"]
        svals   = [s["total_stops"],s["visited"],s["skipped"],
                   str(s["compliance_pct"])+"%",
                   "$"+str(int(s["total_savings"])),"$"+str(int(s["total_losses"])),
                   "$"+str(int(s["net_savings"])),s["flag_count"]]
        scolors = [WHITE,GREEN,RED if s["skipped"] else WHITE,
                   GREEN if s["compliance_pct"]>=80 else (AMBER if s["compliance_pct"]>=60 else RED),
                   GREEN,RED if s["total_losses"] else WHITE,
                   GREEN if s["net_savings"]>=0 else RED,
                   RED if s["flag_count"] else WHITE]
        for ci,(lbl,val,col) in enumerate(zip(slabels,svals,scolors),1):
            cl=ws.cell(row=2,column=ci,value=lbl)
            cl.font=Font(name="Arial",size=9,color=MUTED); cl.fill=_fill(CARD2); cl.alignment=_align("center")
            cv=ws.cell(row=3,column=ci,value=val)
            cv.font=Font(name="Arial",bold=True,color=col,size=12)
            cv.fill=_fill(DARK); cv.alignment=_align("center")
            ws.column_dimensions[get_column_letter(ci)].width=13
        ws.row_dimensions[2].height=18; ws.row_dimensions[3].height=28

        # Fuel Stops
        _sec(ws,5,"Fuel Stops - Every Stop This Week",8)
        sc=[("Date",16),("Station Name",30),("State",6),("Recommended Stop",28),
            ("Fuel %",8),("Card $/gal",12),("Savings ($)",12),("Followed?",10)]
        for ci,(h,w) in enumerate(sc,1): _hdr(ws,6,ci,h,w=w,color=GREEN)
        ws.row_dimensions[6].height=28
        for ri,stop in enumerate(td["fuel_stops"],7):
            fc=GREEN if stop["followed"] else RED
            _cell(ws,ri,1,stop["date"]); _cell(ws,ri,2,stop["station"])
            _cell(ws,ri,3,stop["state"],center=True); _cell(ws,ri,4,stop["recommended"])
            _cell(ws,ri,5,str(int(stop["fuel_pct"]))+"%",
                  center=True,color=AMBER if stop["fuel_pct"]<30 else WHITE)
            _cell(ws,ri,6,stop["card_price"] if stop["card_price"] else 0,fmt="$#,##0.000",center=True)
            _cell(ws,ri,7,stop["savings"],fmt="$#,##0.00",center=True,color=GREEN if stop["savings"]>0 else WHITE)
            _cell(ws,ri,8,"Yes" if stop["followed"] else "No",color=fc,bold=True,center=True)
            ws.row_dimensions[ri].height=20
        last_s=6+len(td["fuel_stops"]); tr=last_s+1
        for ci,v in enumerate(["TOTAL","","","","","",
            "=SUM(G7:G"+str(last_s)+")",""],1):
            c=ws.cell(row=tr,column=ci,value=v)
            c.font=Font(name="Arial",bold=True,color=WHITE,size=10)
            c.fill=_fill(GDRK); c.alignment=_align("center"); c.border=_border()
            if ci==7 and v: c.number_format="$#,##0.00"
        ws.row_dimensions[tr].height=22

        # Flags
        fr=tr+2; _sec(ws,fr,"Flags and Violations",6,color=RED if td["flags"] else GREEN)
        if not td["flags"]:
            nr=fr+1
            ws.merge_cells(start_row=nr,start_column=1,end_row=nr,end_column=6)
            c=ws.cell(row=nr,column=1,value="No flags this week - perfect compliance")
            c.font=Font(name="Arial",color=GREEN,size=10)
            c.fill=_fill(DARK); c.alignment=_align("left")
            ws.row_dimensions[nr].height=20
        else:
            fhr=fr+1
            fc_cols=[("Date",14),("Flag Type",18),("Recommended Stop",30),
                     ("Actual Issue",30),("Fuel %",8),("Savings Lost",14)]
            for ci,(h,w) in enumerate(fc_cols,1): _hdr(ws,fhr,ci,h,w=w,color=RED,bg="3A1010")
            ws.row_dimensions[fhr].height=28
            for ri,flag in enumerate(td["flags"],fhr+1):
                fc2=FLAG_COLORS.get(flag["type"],WHITE)
                _cell(ws,ri,1,flag["date"]); _cell(ws,ri,2,flag["type"],color=fc2,bold=True)
                _cell(ws,ri,3,flag["recommended"]); _cell(ws,ri,4,flag["actual"],color=RED)
                _cell(ws,ri,5,str(flag["fuel_pct"])+"%",center=True,
                      color=AMBER if flag["fuel_pct"]<30 else WHITE)
                _cell(ws,ri,6,flag["loss"] if flag["loss"] else 0,
                      fmt="$#,##0.00",color=RED if flag["loss"]>0 else WHITE,
                      bold=flag["loss"]>0,center=True)
                ws.row_dimensions[ri].height=20

    wb.save(output_path)
    import logging
    logging.getLogger(__name__).info(f"Truck report saved: {output_path} ({len(trucks)} trucks)")
